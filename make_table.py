"""Make table based on results on excel sheets"""

# %% Import libraries
import openpyxl
# Define custom functions


def load_workbook(file_path):
    """Convert noetbook to dictionary"""

    workbook = openpyxl.load_workbook(file_path)
    sheet_names = workbook.sheetnames

    data = {}

    for metric in sheet_names:
        rows = workbook[metric].rows
        condition = ''
        for row in rows:
            first_cell = str(row[0].value)
            # If anova is found effect in conditions or just methods
            if 'Comparisons' in first_cell:
                # Check for condition
                if 'within' in first_cell:
                    condition = first_cell.split(' ')[-1]
                else:
                    condition = 'all'

            # Only check for pvalues in the annotated comparisons
            elif 'Annotated' in first_cell:
                print([cell.value for cell in row])
                method = first_cell.split(' ')[-1]

                # Correct for cases when annotated is the last word
                if method == 'Annotated':
                    method = first_cell.split(' ')[0]
                pvalue = str(row[-2].value)
                # Correct p-values of 1000 to 1
                if pvalue == '1000':
                    pvalue = '1'
                elif pvalue == '<0.001':
                    pvalue = r'\textless0.001***'
                elif float(pvalue) < 0.01:
                    pvalue = f'{pvalue}**'
                elif float(pvalue) < 0.05:
                    pvalue = f'{pvalue}*'

                # Correct method names
                if 'wavelet' in method.lower():
                    method = 'Wavelet_transform'
                elif 'elgendi_et_al' in method.lower():
                    method = 'Elgendi'

                # Correct condition names
                if 'math' in condition:
                    condition = 'math'
                elif 'hand_bike' in condition:
                    condition = 'handbike'

                # Print if p-value is significant
                print(f'{method} - {condition} - {pvalue}')
                data[(metric, method, condition)] = str(pvalue)

    return data


def convert_table(table_name, data):

    tableIO = open(f'{table_name}', 'r', encoding='utf-8')
    newTableIO = open(f'{table_name[:-4]}_p.txt', 'w', encoding='utf-8')

    # Only print lines after METHOD and before \end{tblr}
    print_table = False
    for line in tableIO:

        if r'\end{tblr}' in line:
            print_table = False

        if not print_table:
            newTableIO.write(line)
        else:
            metric, condition = line.replace(
                ' ', '').replace('\\', '').split('&')[0:2]
            nw_line = [metric, condition]

            for method in methods:
                nw_line.append(data[('HRV_'+metric, method, condition)])

            # Write new line
            newTableIO.write(' & '.join(nw_line) + r'\\' + '\n')

        if 'METHOD' in line:
            print_table = True
            methods = line[:-2].replace(' ',
                                        '').replace('\\', '').split('&')[2:]

    tableIO.close()
    newTableIO.close()


# %%
if __name__ == '__main__':
    # %% Load Libraries
    import os

    # %% Look for tables
    tables = os.listdir('tables/')
    tables = list(filter(lambda x: '.txt' in x, tables))

    # %% Load workbooks
    data_cs = load_workbook('tables/estadistica_chest_strap.xlsx')
    data_loose = load_workbook('tables/estadistica_einthoven.xlsx')

    # %% Load workbook
    for table in tables:
        if 'cs' in table:
            convert_table('tables/'+table, data_cs)
        else:
            convert_table('tables/'+table, data_loose)
